#!/usr/bin/env python3
"""
Quét tax_files, liệt kê TẤT CẢ hóa đơn đã bị thay bằng bia/rượu (chỉ có Sapporo, Tiger Draught, Coke).
"""

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
TAX_DIR = BASE_DIR / "tax_files"

# Tên món trong hóa đơn bia/rượu do daily_beverage_invoices tạo
BEVERAGE_KEYWORDS = ("sapporo", "tiger draught", "coke")


def is_beverage_only_invoice(filepath):
    """Kiểm tra file có phải hóa đơn chỉ toàn bia/rượu (đã thay thế) không."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        product_names = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=3).value  # Ten_san_pham
            if val and str(val).strip():
                product_names.append(str(val).strip().lower())
        wb.close()
    except Exception:
        return False, []

    if not product_names:
        return False, []

    for name in product_names:
        if "phí dịch vụ" in name or "service" in name:
            continue
        if not any(kw in name for kw in BEVERAGE_KEYWORDS):
            return False, product_names  # có món không phải bia/rượu
    return True, product_names


def main():
    if not TAX_DIR.exists():
        print("Không tìm thấy thư mục tax_files")
        return 1

    files = sorted(f for f in TAX_DIR.glob("*.xlsx") if not f.name.startswith("Grab - "))
    beverage_invoices = []

    for f in files:
        is_beverage, names = is_beverage_only_invoice(f)
        if is_beverage:
            stem = f.stem  # "070092 - atm - 612.800đ"
            parts = stem.split(" - ")
            inv_id = parts[0].strip() if parts else ""
            payment = parts[1].strip() if len(parts) > 1 else ""
            total_str = parts[2].strip() if len(parts) > 2 else ""
            beverage_invoices.append({
                "file": f.name,
                "invoice_id": inv_id,
                "payment": payment,
                "total_str": total_str,
                "items": names,
            })

    # In báo cáo
    print("=" * 70)
    print("DANH SÁCH TẤT CẢ HÓA ĐƠN ĐÃ BỊ THAY BẰNG BIA/RƯỢU (Sapporo, Tiger Draught, Coke)")
    print("=" * 70)
    print(f"Tổng số hóa đơn trong tax_files (bỏ Grab): {len(files)}")
    print(f"Tổng số hóa đơn bia/rượu (đã thay thế):    {len(beverage_invoices)}")
    print()
    if not beverage_invoices:
        print("Không có hóa đơn nào chỉ toàn bia/rượu.")
        return 0

    print(f"{'#':<4} {'Mã HĐ':<10} {'Thanh toán':<10} {'Tổng tiền':<20} {'File'}")
    print("-" * 70)
    for i, inv in enumerate(beverage_invoices, 1):
        print(f"{i:<4} {inv['invoice_id']:<10} {inv['payment']:<10} {inv['total_str']:<20} {inv['file']}")
    print("=" * 70)
    return 0


if __name__ == "__main__":
    sys.exit(main())
